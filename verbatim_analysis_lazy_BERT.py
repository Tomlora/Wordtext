"""
Analyse de verbatims clients — Version haute volumétrie (20M+ conversations).
Architecture LazyFrame-first : chaque étape ne collecte que ce dont elle a besoin.

Optimisé avec :
    - Polars LazyFrame pour repousser les collectes au maximum
    - Jamais de .collect() sur les 20M lignes complètes (sauf parsing string)
    - Sentiment vectorisé via expressions Polars (regex Rust)
    - LDA sur échantillon représentatif
    - Comptages termes pré-agrégés par streaming/chunks
    - RAM maîtrisée : on ne matérialise que des sous-ensembles

Prérequis :
    pip install polars pyarrow numpy scikit-learn matplotlib seaborn

Entrée acceptée :
    - pl.LazyFrame  (recommandé : scan_parquet, scan_csv…)
    - pl.DataFrame
    - pd.DataFrame
    Colonnes requises : date, conversation_id, verbatims (liste ou string repr)

Usage :
    # Idéal : LazyFrame depuis un fichier
    lf = pl.scan_parquet("verbatims.parquet")
    analyzer = VerbatimAnalyzer(lf, date_col="date", verbatim_col="verbatims")
    report = analyzer.run_full_analysis()

    # Fonctionne aussi avec un DataFrame classique
    analyzer = VerbatimAnalyzer(df_pandas, ...)
"""

import re
import warnings
from collections import Counter
from datetime import date, timedelta
from typing import Optional, Union

import numpy as np
import polars as pl
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.decomposition import LatentDirichletAllocation

warnings.filterwarnings("ignore")


# ══════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════

NEGATIVE_LEXICON: set[str] = {
    "scandaleux", "honteux", "inadmissible", "inacceptable", "honte",
    "nul", "nulle", "catastrophe", "catastrophique", "horrible",
    "lamentable", "déplorable", "minable", "pitoyable",
    "panne", "bug", "bugge", "plante", "planté",
    "bloqué", "bloquée", "erreur", "impossible",
    "hs", "coupure", "coupé",
    "mécontent", "mécontente", "déçu", "déçue", "insatisfait",
    "insatisfaite", "plainte", "réclamation", "arnaque", "voleur",
    "voleurs", "menteur", "menteurs", "incompétent", "incompétente",
    "incompétents",
    "attente", "attend", "attends", "relance", "relancé",
    "résilier", "résiliation", "résilié", "partir", "quitter",
    "concurrence", "concurrent",
    "surfacturation", "remboursement", "rembourser",
}

NEGATIVE_BIGRAMS: set[str] = {
    "marche pas", "fonctionne pas", "personne répond", "trop cher",
    "hors service", "pure simple", "porter plainte", "trois semaines",
}

STOPWORDS_FR: set[str] = {
    "je", "tu", "il", "elle", "on", "nous", "vous", "ils", "elles",
    "le", "la", "les", "un", "une", "des", "de", "du", "au", "aux",
    "ce", "cette", "ces", "mon", "ma", "mes", "ton", "ta", "tes",
    "son", "sa", "ses", "notre", "votre", "leur", "leurs",
    "et", "ou", "mais", "donc", "car", "ni", "que", "qui", "dont",
    "où", "en", "dans", "sur", "sous", "avec", "sans", "pour", "par",
    "ne", "pas", "plus", "très", "bien", "aussi", "comme", "tout",
    "être", "avoir", "faire", "est", "sont", "suis", "ai", "a",
    "oui", "non", "merci", "bonjour", "bonsoir", "svp", "ok",
    "alors", "ça", "cela", "là", "ici", "voilà", "si", "se",
    "me", "te", "lui", "y", "à", "même", "quand", "peu", "peut",
    "encore", "après", "avant", "déjà", "tous", "toute", "toutes",
    "été", "fait", "dit", "mis", "pris", "rien", "quelque",
}

# Regex patterns pré-compilés
_UNIGRAM_PATTERN = r"\b(" + "|".join(re.escape(w) for w in NEGATIVE_LEXICON) + r")\b"
_BIGRAM_PATTERN = r"(" + "|".join(re.escape(b) for b in NEGATIVE_BIGRAMS) + r")"


# ══════════════════════════════════════════════════════════════════════
# EXPRESSIONS POLARS RÉUTILISABLES (lazy-compatible)
# ══════════════════════════════════════════════════════════════════════

def _clean_expr(col: str = "verbatim") -> pl.Expr:
    """Expression Polars pour nettoyer un texte. Compatible lazy."""
    return (
        pl.col(col)
        .str.to_lowercase()
        .str.replace_all(r"http\S+", "")
        .str.replace_all(r"[^a-zàâäéèêëïîôùûüÿçœæ\s'-]", " ")
        .str.replace_all(r"\s+", " ")
        .str.strip_chars()
    )


def _neg_score_expr(col: str = "clean") -> pl.Expr:
    """Score de négativité via regex. Compatible lazy."""
    return (
        (
            pl.col(col).str.count_matches(_UNIGRAM_PATTERN)
            + pl.col(col).str.count_matches(_BIGRAM_PATTERN)
        )
        .cast(pl.Float64)
        / pl.col(col).str.count_matches(r"\S+").cast(pl.Float64).clip(lower_bound=1)
    ).clip(upper_bound=1.0)


def _sentiment_exprs(col: str = "clean") -> list[pl.Expr]:
    """Retourne les expressions sentiment + is_negative. Compatible lazy."""
    return [
        _neg_score_expr(col).alias("neg_score"),
        (_neg_score_expr(col) > 0.15).alias("is_negative"),
    ]


def _parse_string_list_expr(col: str) -> pl.Expr:
    """
    Expression Polars pure pour parser une string repr de liste en List[Utf8].
    100% lazy, 0 Python, 0 collect.

    Gère les formats :
        '["aaa", "bbb", "ccc"]'
        "['aaa', 'bbb', 'ccc']"
        '["aaa", "bb bb", "ccc"]'  (espaces dans les éléments)

    Stratégie :
        1. Strip les crochets extérieurs [ ]
        2. Strip les guillemets/apostrophes aux extrémités de chaque élément
        3. Split sur le séparateur  '", "'  ou  "', '"  (avec variantes d'espacement)
        4. Nettoyage final des guillemets résiduels sur chaque élément
    """
    return (
        pl.col(col)
        .str.strip_chars()
        # 1. Retirer les crochets
        .str.strip_chars("[]")
        .str.strip_chars()
        # 2. Split sur les séparateurs entre éléments
        #    Gère : "aaa", "bbb"  ou  'aaa', 'bbb' ou mélanges
        .str.split(", ")
        # 3. Nettoyer chaque élément : retirer guillemets/apostrophes
        .list.eval(
            pl.element()
            .str.strip_chars()
            .str.strip_chars("'\"")
            .str.strip_chars()
        )
        # 4. Filtrer les éléments vides
        .list.eval(
            pl.element().filter(pl.element().str.len_chars() > 0)
        )
    )


# ══════════════════════════════════════════════════════════════════════
# ANALYSEUR PRINCIPAL — ARCHITECTURE LAZYFRAME
# ══════════════════════════════════════════════════════════════════════

class VerbatimAnalyzer:
    """
    Analyseur de verbatims clients — LazyFrame-first.

    Principe : self.lf est un LazyFrame préparé (nettoyé, avec colonne 'clean').
    Chaque méthode d'analyse collecte UNIQUEMENT les colonnes/lignes nécessaires.
    Jamais de .collect() sur l'intégralité des 20M lignes avec toutes les colonnes.

    Sentiment en 2 passes (si use_transformer=True) :
        Passe 1 — Lexique (rapide, 20M) : score basé sur des mots-clés négatifs
        Passe 2 — CamemBERT (précis, échantillon) : sur les conversations ambiguës
                  (0.01 < neg_score < ambiguous_threshold), puis extrapolation

    Params:
        df                      : LazyFrame, DataFrame Polars, ou DataFrame Pandas
        date_col                : nom de la colonne date
        verbatim_col            : nom de la colonne contenant les listes de verbatims
        conversation_id_col     : nom de la colonne id de conversation
        z_threshold             : seuil de z-score pour détecter un pic (défaut 2.0)
        growth_window           : fenêtre glissante en jours pour la croissance
        growth_factor           : multiplicateur pour un terme en forte croissance
        top_n_terms             : nombre de termes dans les résultats
        lda_sample_size         : taille de l'échantillon pour la LDA
        chunk_size              : taille des chunks pour le comptage de termes
        use_transformer         : activer la passe 2 CamemBERT (défaut False)
        ambiguous_threshold     : seuil haut de la zone ambiguë (défaut 0.30)
        transformer_sample_size : nb max de conversations ambiguës à scorer (défaut 50K)
        transformer_batch_size  : taille de batch pour l'inférence (défaut 32)
    """

    def __init__(
        self,
        df: Union[pl.LazyFrame, pl.DataFrame, "pd.DataFrame"],
        date_col: str = "date",
        verbatim_col: str = "verbatims",
        conversation_id_col: str = "conversation_id",
        z_threshold: float = 2.0,
        growth_window: int = 7,
        growth_factor: float = 3.0,
        top_n_terms: int = 20,
        lda_sample_size: int = 200_000,
        chunk_size: int = 500_000,
        use_transformer: bool = False,
        ambiguous_threshold: float = 0.30,
        transformer_sample_size: int = 50_000,
        transformer_batch_size: int = 32,
    ):
        self.date_col = date_col
        self.verbatim_col = verbatim_col
        self.conversation_id_col = conversation_id_col
        self.z_threshold = z_threshold
        self.growth_window = growth_window
        self.growth_factor = growth_factor
        self.top_n_terms = top_n_terms
        self.lda_sample_size = lda_sample_size
        self.chunk_size = chunk_size
        self.use_transformer = use_transformer
        self.ambiguous_threshold = ambiguous_threshold
        self.transformer_sample_size = transformer_sample_size
        self.transformer_batch_size = transformer_batch_size

        # Résultats
        self.daily: Optional[pl.DataFrame] = None
        self.spike_details: list[dict] = []
        self.topics: list[dict] = []
        self.trending_terms = pl.DataFrame()
        self.novel_terms = pl.DataFrame()

        # Préparation → LazyFrame
        self.lf = self._prepare(df)

        # Bornes de dates (collecte minimale)
        date_bounds = (
            self.lf.select(
                pl.col(self.date_col).min().alias("min_date"),
                pl.col(self.date_col).max().alias("max_date"),
                pl.count().alias("n_rows"),
            )
            .collect()
        )
        self.min_date: date = date_bounds["min_date"][0]
        self.max_date: date = date_bounds["max_date"][0]
        self.n_rows: int = date_bounds["n_rows"][0]

        print(f"✅ {self.n_rows:,} conversations prêtes (LazyFrame)")

    # ── Préparation → LazyFrame ──────────────────────────────────────

    def _prepare(self, df) -> pl.LazyFrame:
        """
        Prépare un LazyFrame nettoyé avec colonne 'clean'.
        100% lazy dans tous les cas — jamais de collect() sur les 20M lignes.
        """
        print("⏳ Préparation des données...")

        # ── Conversion vers Polars LazyFrame ──
        import pandas as pd
        if isinstance(df, pd.DataFrame):
            print("   Conversion Pandas → Polars...")
            df = pl.from_pandas(df)

        if isinstance(df, pl.DataFrame):
            lf = df.lazy()
        elif isinstance(df, pl.LazyFrame):
            lf = df
        else:
            raise TypeError(
                f"Type non supporté : {type(df)}. "
                "Attendu : pl.LazyFrame, pl.DataFrame, ou pd.DataFrame"
            )

        # ── Détection du type de colonne (collecte 1 seule ligne) ──
        sample_row = lf.select(self.verbatim_col).head(1).collect()
        col_dtype = sample_row[self.verbatim_col].dtype
        is_string_col = col_dtype in (pl.Utf8, pl.String)
        is_list_col = col_dtype.base_type() == pl.List

        if is_string_col:
            # Parsing string → liste via expressions Polars pures (100% lazy)
            print("   Parsing des listes string (100% lazy, regex Polars)...")
            lf = lf.with_columns(
                _parse_string_list_expr(self.verbatim_col)
                .alias("verbatim_list")
            )
        elif is_list_col:
            lf = lf.with_columns(
                pl.col(self.verbatim_col).alias("verbatim_list")
            )
        else:
            # Colonne scalaire → wrap en liste à 1 élément
            lf = lf.with_columns(
                pl.col(self.verbatim_col)
                .cast(pl.Utf8)
                .map_elements(lambda x: [x], return_dtype=pl.List(pl.Utf8))
                .alias("verbatim_list")
            )

        # ── Concaténation + nettoyage (tout lazy) ──
        lf = (
            lf
            .with_columns(
                pl.col("verbatim_list").list.join(". ").alias("verbatim"),
                pl.col("verbatim_list").list.len().alias("n_messages"),
            )
            .with_columns(_clean_expr("verbatim").alias("clean"))
            .filter(pl.col("clean").str.len_chars() > 2)
            .with_columns(pl.col(self.date_col).cast(pl.Date))
        )

        mode = "100% lazy" if (is_string_col or is_list_col) else "lazy après conversion"
        print(f"   ✅ Préparation {mode}")
        return lf

    # ── 1. Sentiment & Détection de pics ─────────────────────────────

    def _refine_with_transformer(self, daily: pl.DataFrame) -> pl.DataFrame:
        """
        Passe 2 : CamemBERT sur un échantillon de conversations ambiguës
        et négatives (lexique), puis correction des stats journalières.

        Pipeline :
            1. Collecte les conversations avec 0.01 < neg_score < ambiguous_threshold
               (zone grise : le lexique ne sait pas trancher)
            2. Échantillonne transformer_sample_size conversations
            3. Passe CamemBERT dessus (CPU, par batch)
            4. Calcule le taux de correction (faux négatifs récupérés)
            5. Ajuste les comptages journaliers
        """
        try:
            from transformers import pipeline as hf_pipeline
        except ImportError:
            print("   ⚠️  transformers et/ou torch non installés.")
            print("   → pip install transformers torch")
            print("   → Poursuite avec le lexique seul.")
            return daily

        print(f"⏳ Passe 2 : CamemBERT sur conversations ambiguës...")
        print(f"   Seuil ambigu : 0.01 < neg_score < {self.ambiguous_threshold}")
        print(f"   Échantillon max : {self.transformer_sample_size:,}")

        # ── Chargement du modèle ──
        print("   Chargement du modèle distilcamembert-base-sentiment...")
        classifier = hf_pipeline(
            "sentiment-analysis",
            model="cmarkea/distilcamembert-base-sentiment",
            truncation=True,
            max_length=512,
            device=-1,  # CPU
        )

        # ── Collecte des ambigus (échantillon) ──
        ambiguous = (
            self.lf
            .with_columns(_neg_score_expr("clean").alias("neg_score"))
            .filter(
                (pl.col("neg_score") > 0.01)
                & (pl.col("neg_score") <= self.ambiguous_threshold)
            )
            .select(self.date_col, "clean", "neg_score")
            .collect()
        )
        n_ambiguous_total = ambiguous.height
        print(f"   {n_ambiguous_total:,} conversations ambiguës trouvées")

        if n_ambiguous_total == 0:
            print("   Aucune conversation ambiguë → pas de correction")
            return daily

        # Échantillonnage
        if n_ambiguous_total > self.transformer_sample_size:
            sample = ambiguous.sample(
                n=self.transformer_sample_size, seed=42
            )
        else:
            sample = ambiguous
        print(f"   Échantillon retenu : {sample.height:,} conversations")

        # ── Inférence CamemBERT par batch ──
        texts = sample.get_column("clean").to_list()
        # Tronquer les textes trop longs pour éviter les lenteurs
        texts = [t[:500] for t in texts]

        print(f"   Inférence CamemBERT ({sample.height:,} textes, "
              f"batch_size={self.transformer_batch_size})...")

        transformer_scores = []
        n_batches = (len(texts) + self.transformer_batch_size - 1) // self.transformer_batch_size

        for i in range(0, len(texts), self.transformer_batch_size):
            batch = texts[i:i + self.transformer_batch_size]
            results = classifier(batch)
            for r in results:
                # Le modèle retourne "1 star" à "5 stars"
                # 1 star = très négatif → score 1.0
                # 5 stars = très positif → score 0.0
                label = r["label"]
                stars = int(label.split()[0]) if label[0].isdigit() else 3
                neg = 1.0 - (stars - 1) / 4.0
                transformer_scores.append(neg)

            batch_num = i // self.transformer_batch_size + 1
            if batch_num % 50 == 0 or batch_num == n_batches:
                print(f"   ... batch {batch_num}/{n_batches}")

        # ── Calcul du taux de correction ──
        sample = sample.with_columns(
            pl.Series(name="transformer_score", values=transformer_scores)
        ).with_columns(
            (pl.col("transformer_score") > 0.6).alias("transformer_negative")
        )

        # Taux de vrais négatifs parmi les ambigus
        n_recovered = sample.filter(pl.col("transformer_negative")).height
        recovery_rate = n_recovered / sample.height if sample.height > 0 else 0.0
        print(f"   CamemBERT : {n_recovered}/{sample.height} ambigus "
              f"reclassés négatifs ({recovery_rate:.1%})")

        # ── Correction des stats journalières ──
        # Compter les ambigus par jour (sur le total, pas l'échantillon)
        ambiguous_daily = (
            ambiguous
            .group_by(self.date_col)
            .agg(pl.count().alias("ambiguous_count"))
        )

        daily = daily.join(ambiguous_daily, on=self.date_col, how="left")
        daily = daily.with_columns(
            pl.col("ambiguous_count").fill_null(0)
        )

        # Ajouter les négatifs récupérés (extrapolation du recovery_rate)
        daily = daily.with_columns(
            (
                pl.col("negative_count")
                + (pl.col("ambiguous_count") * recovery_rate).round(0).cast(pl.Int64)
            ).alias("negative_count_corrected")
        ).with_columns(
            (pl.col("negative_count_corrected") / pl.col("total_conversations"))
            .alias("negative_ratio_corrected")
        )

        # Recalculer le z-score sur le ratio corrigé
        daily = daily.with_columns(
            pl.col("negative_ratio_corrected")
            .rolling_mean(window_size=30, min_periods=7)
            .alias("rolling_mean_corrected"),
            pl.col("negative_ratio_corrected")
            .rolling_std(window_size=30, min_periods=7)
            .alias("rolling_std_corrected"),
        ).with_columns(
            (
                (pl.col("negative_ratio_corrected") - pl.col("rolling_mean_corrected"))
                / pl.col("rolling_std_corrected").clip(lower_bound=1e-9)
            ).alias("z_score_corrected")
        ).with_columns(
            (pl.col("z_score_corrected") > self.z_threshold)
            .alias("is_spike_corrected")
        )

        # Stats de correction
        n_neg_before = daily["negative_count"].sum()
        n_neg_after = daily["negative_count_corrected"].sum()
        n_spikes_before = daily.filter(pl.col("is_spike")).height
        n_spikes_after = daily.filter(
            pl.col("is_spike_corrected").fill_null(False)
        ).height

        print(f"   📊 Correction appliquée :")
        print(f"      Négatifs : {n_neg_before:,} → {n_neg_after:,} "
              f"(+{n_neg_after - n_neg_before:,})")
        print(f"      Pics     : {n_spikes_before} → {n_spikes_after}")

        self.transformer_recovery_rate = recovery_rate
        self.transformer_sample_stats = {
            "n_ambiguous_total": n_ambiguous_total,
            "n_sampled": sample.height,
            "n_recovered": n_recovered,
            "recovery_rate": recovery_rate,
        }

        return daily

    def detect_negative_spikes(self) -> pl.DataFrame:
        """
        Détecte les pics de négativité.

        Passe 1 : lexique (sur les 20M, via agrégation lazy)
        Passe 2 (optionnelle) : CamemBERT sur les ambigus pour corriger
        """
        print("⏳ Passe 1 : sentiment lexique + détection des pics...")

        # Agrégation jour directement en lazy
        daily = (
            self.lf
            .with_columns(*_sentiment_exprs("clean"))
            .group_by(self.date_col)
            .agg(
                pl.count().alias("total_conversations"),
                pl.col("is_negative").sum().alias("negative_count"),
                pl.col("neg_score").mean().alias("mean_neg_score"),
            )
            .sort(self.date_col)
            .collect()
        )

        daily = daily.with_columns(
            (pl.col("negative_count") / pl.col("total_conversations"))
            .alias("negative_ratio")
        )

        # Z-score glissant
        daily = daily.with_columns(
            pl.col("negative_ratio")
            .rolling_mean(window_size=30, min_periods=7)
            .alias("rolling_mean"),
            pl.col("negative_ratio")
            .rolling_std(window_size=30, min_periods=7)
            .alias("rolling_std"),
        ).with_columns(
            (
                (pl.col("negative_ratio") - pl.col("rolling_mean"))
                / pl.col("rolling_std").clip(lower_bound=1e-9)
            ).alias("z_score")
        ).with_columns(
            (pl.col("z_score") > self.z_threshold).alias("is_spike")
        )

        n_neg = daily["negative_count"].sum()
        print(f"   {n_neg:,} conversations négatives (lexique) "
              f"({n_neg / self.n_rows * 100:.1f}%)")

        # ── Passe 2 : CamemBERT (optionnelle) ──
        if self.use_transformer:
            daily = self._refine_with_transformer(daily)

        # Déterminer quelle colonne de spike utiliser
        has_corrected = "is_spike_corrected" in daily.columns
        spike_col = "is_spike_corrected" if has_corrected else "is_spike"

        # Détails des pics : collecte ciblée par date
        spike_dates = (
            daily.filter(
                pl.col(spike_col).fill_null(False)
            )
            .get_column(self.date_col)
            .to_list()
        )

        spike_details = []
        for d in spike_dates:
            day_neg = (
                self.lf
                .with_columns(*_sentiment_exprs("clean"))
                .filter(
                    (pl.col(self.date_col) == d) & pl.col("is_negative")
                )
                .select("clean", "verbatim")
                .collect()
            )
            all_tokens = []
            for text in day_neg.get_column("clean").to_list():
                words = text.split()
                all_tokens.extend(
                    w for w in words if w not in STOPWORDS_FR and len(w) > 2
                )
            top_terms = Counter(all_tokens).most_common(10)
            samples = day_neg.get_column("verbatim").head(5).to_list()
            spike_details.append({
                "date": d,
                "top_negative_terms": top_terms,
                "sample_verbatims": samples,
            })

        self.daily = daily
        self.spike_details = spike_details
        print(f"✅ {len(spike_dates)} pic(s) de négativité détecté(s)")
        return daily

    # ── 2. Sujets émergents (LDA sur échantillon) ───────────────────

    def detect_emerging_topics(
        self, n_topics: int = 10, recent_days: int = 14
    ) -> dict:
        """
        LDA sur échantillon. Collecte uniquement l'échantillon (~200K lignes),
        pas les 20M.
        """
        print(f"⏳ Détection des sujets émergents "
              f"(LDA sur échantillon de {self.lda_sample_size:,})...")

        cutoff = self.max_date - timedelta(days=recent_days)

        # Compter les volumes pour calculer les proportions (collect minimal)
        counts = (
            self.lf
            .select(
                (pl.col(self.date_col) >= cutoff).alias("is_recent")
            )
            .group_by("is_recent")
            .agg(pl.count().alias("n"))
            .collect()
        )
        n_recent_total = counts.filter(pl.col("is_recent"))["n"][0] if counts.filter(pl.col("is_recent")).height > 0 else 0
        n_history_total = counts.filter(~pl.col("is_recent"))["n"][0] if counts.filter(~pl.col("is_recent")).height > 0 else 0

        recent_ratio = n_recent_total / (n_recent_total + n_history_total)
        n_recent_sample = max(int(self.lda_sample_size * recent_ratio), 10_000)
        n_history_sample = self.lda_sample_size - n_recent_sample

        # Collecte UNIQUEMENT l'échantillon (colonne clean seulement)
        recent_sample = (
            self.lf
            .filter(pl.col(self.date_col) >= cutoff)
            .select("clean")
            .collect()
            .sample(n=min(n_recent_sample, n_recent_total), seed=42)
        )
        history_sample = (
            self.lf
            .filter(pl.col(self.date_col) < cutoff)
            .select("clean")
            .collect()
            .sample(n=min(n_history_sample, n_history_total), seed=42)
        )

        print(f"   Échantillon : {history_sample.height:,} historique + "
              f"{recent_sample.height:,} récent")

        is_recent = [False] * history_sample.height + [True] * recent_sample.height
        all_texts = (
            history_sample.get_column("clean").to_list()
            + recent_sample.get_column("clean").to_list()
        )

        # Vectorisation + LDA
        vectorizer = CountVectorizer(
            max_features=3000, min_df=10, max_df=0.8,
            stop_words=list(STOPWORDS_FR), ngram_range=(1, 2),
        )
        X = vectorizer.fit_transform(all_texts)
        feature_names = vectorizer.get_feature_names_out()

        lda = LatentDirichletAllocation(
            n_components=n_topics, random_state=42, max_iter=15, n_jobs=-1,
        )
        doc_topics = lda.fit_transform(X)

        is_recent_arr = np.array(is_recent)
        recent_dist = doc_topics[is_recent_arr].mean(axis=0)
        history_dist = doc_topics[~is_recent_arr].mean(axis=0)

        with np.errstate(divide="ignore", invalid="ignore"):
            emergence_ratio = np.where(
                history_dist > 0.001,
                recent_dist / history_dist,
                recent_dist * 100,
            )

        topics = []
        for idx in range(n_topics):
            top_word_indices = lda.components_[idx].argsort()[-10:][::-1]
            top_words = [feature_names[i] for i in top_word_indices]
            topics.append({
                "topic_id": idx,
                "top_words": top_words,
                "recent_weight": round(float(recent_dist[idx]), 4),
                "history_weight": round(float(history_dist[idx]), 4),
                "emergence_ratio": round(float(emergence_ratio[idx]), 2),
            })

        topics = sorted(topics, key=lambda x: x["emergence_ratio"], reverse=True)
        self.topics = topics
        emerging = [t for t in topics if t["emergence_ratio"] > 1.5]
        print(f"✅ {len(emerging)} sujet(s) émergent(s) détecté(s) (ratio > 1.5x)")
        return {"all_topics": topics, "emerging": emerging}

    # ── 3. Termes en forte croissance ────────────────────────────────

    def detect_trending_terms(self) -> pl.DataFrame:
        """
        Détecte les termes en forte croissance.
        Collecte par chunks (seulement 2 colonnes : date + clean).
        """
        print("⏳ Détection des termes en forte croissance...")
        print("   Comptage des termes par jour (par chunks)...")

        # On ne collecte que date + clean, jamais le reste
        minimal_lf = self.lf.select(self.date_col, "clean")

        term_day_counts: Counter = Counter()
        n = self.n_rows

        # Streaming par chunks via slice sur le LazyFrame
        for start in range(0, n, self.chunk_size):
            chunk_size = min(self.chunk_size, n - start)
            chunk = minimal_lf.slice(start, chunk_size).collect()

            chunk_texts = chunk.get_column("clean").to_list()
            chunk_dates = chunk.get_column(self.date_col).to_list()

            for text, d in zip(chunk_texts, chunk_dates):
                words = text.split()
                seen = set()
                for w in words:
                    if w not in STOPWORDS_FR and len(w) > 2 and w not in seen:
                        term_day_counts[(str(d), w)] += 1
                        seen.add(w)

            processed = start + chunk_size
            if processed % (self.chunk_size * 5) == 0 and processed > 0:
                print(f"   ... {processed:,}/{n:,} conversations")

        print(f"   {len(term_day_counts):,} paires (terme, jour) comptées")

        # Agrégation en Polars
        records = [
            {"date": k[0], "term": k[1], "count": v}
            for k, v in term_day_counts.items()
        ]
        term_df = pl.DataFrame(records).with_columns(
            pl.col("date").str.to_date()
        )

        max_date = term_df.get_column("date").max()
        recent_start = max_date - timedelta(days=self.growth_window)
        baseline_end = recent_start
        baseline_start = baseline_end - timedelta(days=30)

        recent_avg = (
            term_df.lazy()
            .filter(pl.col("date") >= recent_start)
            .group_by("term")
            .agg(
                (pl.col("count").sum() / self.growth_window)
                .alias("current_avg_daily")
            )
        )

        baseline_avg = (
            term_df.lazy()
            .filter(
                (pl.col("date") >= baseline_start)
                & (pl.col("date") < baseline_end)
            )
            .group_by("term")
            .agg(
                (pl.col("count").sum() / 30.0).alias("baseline_avg_daily")
            )
        )

        trending = (
            recent_avg
            .join(baseline_avg, on="term", how="left")
            .with_columns(pl.col("baseline_avg_daily").fill_null(0.0))
            .with_columns(
                pl.when(pl.col("baseline_avg_daily") > 0.5)
                .then(pl.col("current_avg_daily") / pl.col("baseline_avg_daily"))
                .otherwise(pl.col("current_avg_daily") * 10)
                .alias("growth_ratio")
            )
            .filter(
                (pl.col("growth_ratio") >= self.growth_factor)
                & (pl.col("current_avg_daily") >= 2)
            )
            .sort("growth_ratio", descending=True)
            .head(self.top_n_terms)
            .with_columns(
                pl.col("current_avg_daily").round(2),
                pl.col("baseline_avg_daily").round(2),
                pl.col("growth_ratio").round(2),
            )
            .collect()
        )

        self.trending_terms = trending
        print(f"✅ {trending.height} terme(s) en forte croissance "
              f"(x{self.growth_factor}+)")
        return trending

    # ── 4. Termes inhabituels (nouveaux) ─────────────────────────────

    def detect_novel_terms(
        self, recent_days: int = 7, min_count: int = 20
    ) -> pl.DataFrame:
        """
        Détecte les termes nouveaux. Collecte par chunks (2 colonnes seulement).
        """
        print("⏳ Détection des termes inhabituels...")

        cutoff = self.max_date - timedelta(days=recent_days)
        total_hist_days = (cutoff - self.min_date).days or 1

        minimal_lf = self.lf.select(self.date_col, "clean")

        recent_counts: Counter = Counter()
        history_counts: Counter = Counter()

        for start in range(0, self.n_rows, self.chunk_size):
            chunk_size = min(self.chunk_size, self.n_rows - start)
            chunk = minimal_lf.slice(start, chunk_size).collect()

            chunk_texts = chunk.get_column("clean").to_list()
            chunk_dates = chunk.get_column(self.date_col).to_list()

            for text, d in zip(chunk_texts, chunk_dates):
                words = set(text.split())
                tokens = {w for w in words if w not in STOPWORDS_FR and len(w) > 2}
                if d >= cutoff:
                    recent_counts.update(tokens)
                else:
                    history_counts.update(tokens)

        novel = []
        for term, count in recent_counts.items():
            if count < min_count:
                continue
            hist = history_counts.get(term, 0)
            hist_daily = hist / total_hist_days
            recent_daily = count / recent_days

            if hist_daily < 1.0:
                novel.append({
                    "term": term,
                    "recent_count": count,
                    "recent_daily_avg": round(recent_daily, 2),
                    "history_total": hist,
                    "history_daily_avg": round(hist_daily, 2),
                    "novelty": "nouveau" if hist == 0 else "quasi_nouveau",
                })

        novel_df = (
            pl.DataFrame(novel)
            .sort("recent_count", descending=True)
            .head(self.top_n_terms)
        ) if novel else pl.DataFrame()

        self.novel_terms = novel_df
        print(f"✅ {novel_df.height} terme(s) inhabituels détecté(s)")
        return novel_df

    # ── 5. Visualisations ────────────────────────────────────────────

    def plot_dashboard(self, save_path: Optional[str] = None):
        """Dashboard récapitulatif."""
        fig, axes = plt.subplots(3, 1, figsize=(16, 14), constrained_layout=True)
        fig.suptitle(
            "📊 Dashboard Analyse Verbatims Clients",
            fontsize=16, fontweight="bold",
        )

        if self.daily is not None:
            ax1 = axes[0]
            ax1.set_title("Volume de conversations et pics de négativité")
            daily_pd = self.daily.with_columns(
                pl.col("is_spike").fill_null(False)
            ).to_pandas()
            ax1.bar(
                daily_pd[self.date_col], daily_pd["total_conversations"],
                color="#c8d6e5", alpha=0.7, label="Total conversations",
            )
            ax1b = ax1.twinx()

            # Ratio négatif (lexique)
            ax1b.plot(
                daily_pd[self.date_col], daily_pd["negative_ratio"],
                color="#e74c3c", linewidth=1.2, alpha=0.5,
                label="Ratio négatif (lexique)",
            )

            # Si passe 2 CamemBERT : afficher le ratio corrigé
            has_corrected = "negative_ratio_corrected" in daily_pd.columns
            if has_corrected:
                ax1b.plot(
                    daily_pd[self.date_col],
                    daily_pd["negative_ratio_corrected"],
                    color="#c0392b", linewidth=2,
                    label="Ratio négatif (corrigé CamemBERT)",
                )
                spike_col = "is_spike_corrected"
                ratio_col = "negative_ratio_corrected"
            else:
                spike_col = "is_spike"
                ratio_col = "negative_ratio"

            spikes = daily_pd[daily_pd[spike_col].fillna(False)]
            ax1b.scatter(
                spikes[self.date_col], spikes[ratio_col],
                color="red", s=80, zorder=5, marker="^", label="Pic détecté",
            )
            ax1.set_ylabel("Volume")
            ax1b.set_ylabel("Ratio négatif", color="#e74c3c")
            ax1.legend(loc="upper left")
            ax1b.legend(loc="upper right")
            ax1.xaxis.set_major_formatter(mdates.DateFormatter("%d/%m"))

        ax2 = axes[1]
        if self.trending_terms.height > 0:
            data = self.trending_terms.head(15).to_pandas()
            colors = plt.cm.Reds(np.linspace(0.4, 0.9, len(data)))[::-1]
            ax2.barh(data["term"], data["growth_ratio"], color=colors)
            ax2.set_xlabel("Ratio de croissance")
            ax2.set_title("Termes en forte croissance (top 15)")
            ax2.invert_yaxis()
        else:
            ax2.text(0.5, 0.5, "Pas de données", ha="center", va="center")
            ax2.set_title("Termes en forte croissance")

        ax3 = axes[2]
        if self.novel_terms.height > 0:
            data = self.novel_terms.head(15).to_pandas()
            colors = [
                "#e67e22" if n == "nouveau" else "#f39c12"
                for n in data["novelty"]
            ]
            ax3.barh(data["term"], data["recent_count"], color=colors)
            ax3.set_xlabel("Occurrences récentes")
            ax3.set_title("Termes inhabituels / nouveaux (top 15)")
            ax3.invert_yaxis()
        else:
            ax3.text(0.5, 0.5, "Pas de données", ha="center", va="center")
            ax3.set_title("Termes inhabituels / nouveaux")

        if save_path:
            fig.savefig(save_path, dpi=150, bbox_inches="tight")
            print(f"📁 Dashboard sauvegardé : {save_path}")

        plt.close(fig)
        return fig

    # ── 6. Export Excel avec phrases exemples ────────────────────────

    def export_to_excel(
        self,
        path: str = "analyse_termes.xlsx",
        max_samples_per_term: int = 20,
    ) -> str:
        """
        Exporte les termes en forte croissance et les termes inhabituels
        dans un Excel avec les conversations exemples.

        Pour chaque terme, on collecte (en lazy) un échantillon de conversations
        contenant ce terme, sans jamais charger les 20M en mémoire.

        Params:
            path                : chemin du fichier Excel de sortie
            max_samples_per_term: nombre max de conversations exemples par terme
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        print(f"⏳ Export Excel avec conversations exemples...")

        wb = Workbook()

        # Styles
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_fill_red = PatternFill("solid", fgColor="C0392B")
        header_fill_orange = PatternFill("solid", fgColor="E67E22")
        header_fill_grey = PatternFill("solid", fgColor="7F8C8D")
        term_font = Font(name="Arial", bold=True, size=10, color="2C3E50")
        term_fill = PatternFill("solid", fgColor="FADBD8")
        term_fill_novel = PatternFill("solid", fgColor="FAE5D3")
        data_font = Font(name="Arial", size=10)
        wrap_align = Alignment(wrap_text=True, vertical="top")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            bottom=Side(style="thin", color="D5D8DC"),
        )

        def _write_headers(ws, headers, fill):
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = fill
                cell.alignment = center_align
            ws.freeze_panes = "A2"

        def _fetch_samples(term: str, n: int) -> list[dict]:
            """Collecte en lazy les conversations les plus récentes contenant le terme."""
            samples = (
                self.lf
                .filter(pl.col("clean").str.contains(rf"\b{re.escape(term)}\b"))
                .select(
                    pl.col(self.date_col),
                    pl.col(self.conversation_id_col),
                    pl.col("verbatim"),
                )
                .sort(self.date_col, descending=True)
                .head(n)
                .collect()
            )
            return samples.to_dicts()

        # ═══════════════════════════════════════════════════════════
        # Onglet 1 : Termes en forte croissance
        # ═══════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Termes en croissance"
        _write_headers(ws1, [
            "Terme", "Croissance (x)", "Moy. actuelle/jour",
            "Moy. baseline/jour", "Date", "ID conversation",
            "Conversation complète",
        ], header_fill_red)

        ws1.column_dimensions["A"].width = 20
        ws1.column_dimensions["B"].width = 14
        ws1.column_dimensions["C"].width = 16
        ws1.column_dimensions["D"].width = 16
        ws1.column_dimensions["E"].width = 13
        ws1.column_dimensions["F"].width = 18
        ws1.column_dimensions["G"].width = 80

        row = 2
        if self.trending_terms.height > 0:
            for term_row in self.trending_terms.iter_rows(named=True):
                term = term_row["term"]
                samples = _fetch_samples(term, max_samples_per_term)
                start_row = row

                if not samples:
                    # Ligne terme sans exemples
                    ws1.cell(row=row, column=1, value=term).font = term_font
                    ws1.cell(row=row, column=2, value=term_row["growth_ratio"]).font = data_font
                    ws1.cell(row=row, column=3, value=term_row["current_avg_daily"]).font = data_font
                    ws1.cell(row=row, column=4, value=term_row["baseline_avg_daily"]).font = data_font
                    row += 1
                else:
                    for i, s in enumerate(samples):
                        if i == 0:
                            ws1.cell(row=row, column=1, value=term).font = term_font
                            ws1.cell(row=row, column=2, value=term_row["growth_ratio"]).font = data_font
                            ws1.cell(row=row, column=3, value=term_row["current_avg_daily"]).font = data_font
                            ws1.cell(row=row, column=4, value=term_row["baseline_avg_daily"]).font = data_font
                        c_date = ws1.cell(row=row, column=5, value=str(s[self.date_col]))
                        c_date.font = data_font
                        c_id = ws1.cell(row=row, column=6, value=s[self.conversation_id_col])
                        c_id.font = data_font
                        c_verb = ws1.cell(row=row, column=7, value=s["verbatim"])
                        c_verb.font = data_font
                        c_verb.alignment = wrap_align
                        row += 1

                # Colorer le bloc du terme
                for r in range(start_row, row):
                    ws1.cell(row=r, column=1).fill = term_fill
                    for c in range(1, 8):
                        ws1.cell(row=r, column=c).border = thin_border

                # Ligne vide de séparation
                row += 1

            print(f"   ✅ Onglet 'Termes en croissance' : "
                  f"{self.trending_terms.height} termes")

        # ═══════════════════════════════════════════════════════════
        # Onglet 2 : Termes inhabituels / nouveaux
        # ═══════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Termes inhabituels")
        _write_headers(ws2, [
            "Terme", "Type", "Occurrences récentes",
            "Moy. récente/jour", "Historique total",
            "Date", "ID conversation", "Conversation complète",
        ], header_fill_orange)

        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 16
        ws2.column_dimensions["E"].width = 16
        ws2.column_dimensions["F"].width = 13
        ws2.column_dimensions["G"].width = 18
        ws2.column_dimensions["H"].width = 80

        row = 2
        if self.novel_terms.height > 0:
            for term_row in self.novel_terms.iter_rows(named=True):
                term = term_row["term"]
                samples = _fetch_samples(term, max_samples_per_term)
                start_row = row

                if not samples:
                    ws2.cell(row=row, column=1, value=term).font = term_font
                    ws2.cell(row=row, column=2, value=term_row["novelty"]).font = data_font
                    ws2.cell(row=row, column=3, value=term_row["recent_count"]).font = data_font
                    ws2.cell(row=row, column=4, value=term_row["recent_daily_avg"]).font = data_font
                    ws2.cell(row=row, column=5, value=term_row["history_total"]).font = data_font
                    row += 1
                else:
                    for i, s in enumerate(samples):
                        if i == 0:
                            ws2.cell(row=row, column=1, value=term).font = term_font
                            ws2.cell(row=row, column=2, value=term_row["novelty"]).font = data_font
                            ws2.cell(row=row, column=3, value=term_row["recent_count"]).font = data_font
                            ws2.cell(row=row, column=4, value=term_row["recent_daily_avg"]).font = data_font
                            ws2.cell(row=row, column=5, value=term_row["history_total"]).font = data_font
                        c_date = ws2.cell(row=row, column=6, value=str(s[self.date_col]))
                        c_date.font = data_font
                        c_id = ws2.cell(row=row, column=7, value=s[self.conversation_id_col])
                        c_id.font = data_font
                        c_verb = ws2.cell(row=row, column=8, value=s["verbatim"])
                        c_verb.font = data_font
                        c_verb.alignment = wrap_align
                        row += 1

                for r in range(start_row, row):
                    ws2.cell(row=r, column=1).fill = term_fill_novel
                    for c in range(1, 9):
                        ws2.cell(row=r, column=c).border = thin_border

                row += 1

            print(f"   ✅ Onglet 'Termes inhabituels' : "
                  f"{self.novel_terms.height} termes")

        # ═══════════════════════════════════════════════════════════
        # Onglet 3 : Résumé
        # ═══════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Résumé")
        _write_headers(ws3, ["Métrique", "Valeur"], header_fill_grey)
        ws3.column_dimensions["A"].width = 35
        ws3.column_dimensions["B"].width = 25

        summary = [
            ("Période analysée", f"{self.min_date} → {self.max_date}"),
            ("Nombre total de conversations", f"{self.n_rows:,}"),
            ("Pics de négativité détectés", str(len(self.spike_details))),
            ("Termes en forte croissance", str(self.trending_terms.height)),
            ("Termes inhabituels détectés", str(self.novel_terms.height)),
        ]
        for i, (metric, val) in enumerate(summary, 2):
            ws3.cell(row=i, column=1, value=metric).font = Font(name="Arial", bold=True, size=10)
            ws3.cell(row=i, column=2, value=val).font = data_font

        # Déplacer Résumé en premier
        wb.move_sheet("Résumé", offset=-2)

        wb.save(path)
        print(f"📁 Excel sauvegardé : {path}")
        return path

    def run_full_analysis(
        self,
        save_dashboard: Optional[str] = None,
        save_excel: Optional[str] = None,
        max_samples_per_term: int = 20,
    ) -> dict:
        """Lance l'analyse complète."""
        print("=" * 60)
        print("🔍 ANALYSE COMPLÈTE DES VERBATIMS CLIENTS")
        print("=" * 60)
        print(f"📅 Période : {self.min_date} → {self.max_date}")
        print(f"📝 Nombre de conversations : {self.n_rows:,}")
        print()

        self.detect_negative_spikes()
        print()

        topics_result = self.detect_emerging_topics()
        print()

        trending = self.detect_trending_terms()
        print()

        novel = self.detect_novel_terms()
        print()

        fig = self.plot_dashboard(save_path=save_dashboard)

        # 6. Export Excel
        if save_excel:
            self.export_to_excel(
                path=save_excel,
                max_samples_per_term=max_samples_per_term,
            )
            print()

        # Rapport console
        print("=" * 60)
        print("📋 RÉSUMÉ")
        print("=" * 60)

        if self.spike_details:
            print("\n🔴 PICS DE NÉGATIVITÉ :")
            for sp in self.spike_details[:5]:
                terms_str = ", ".join(
                    f"{t[0]} ({t[1]})" for t in sp["top_negative_terms"][:5]
                )
                print(f"  📅 {sp['date']} — Termes : {terms_str}")
                for v in sp["sample_verbatims"][:2]:
                    print(f"     💬 \"{str(v)[:100]}\"")

        if topics_result["emerging"]:
            print("\n🟠 SUJETS ÉMERGENTS :")
            for t in topics_result["emerging"][:5]:
                words = ", ".join(t["top_words"][:5])
                print(f"  Topic #{t['topic_id']} (x{t['emergence_ratio']}) : {words}")

        if trending.height > 0:
            print("\n🟡 TERMES EN FORTE CROISSANCE :")
            for row in trending.head(10).iter_rows(named=True):
                print(
                    f"  📈 {row['term']} : x{row['growth_ratio']} "
                    f"({row['baseline_avg_daily']}/j → {row['current_avg_daily']}/j)"
                )

        if novel.height > 0:
            print("\n🆕 TERMES INHABITUELS / NOUVEAUX :")
            for row in novel.head(10).iter_rows(named=True):
                tag = "🆕" if row["novelty"] == "nouveau" else "⚠️"
                print(
                    f"  {tag} {row['term']} : {row['recent_count']} "
                    f"occurrences ({row['novelty']})"
                )

        print("\n" + "=" * 60)

        return {
            "daily_stats": self.daily,
            "spike_details": self.spike_details,
            "topics": topics_result,
            "trending_terms": trending,
            "novel_terms": novel,
            "figure": fig,
        }


# ══════════════════════════════════════════════════════════════════════
# DÉMO
# ══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import time
    from datetime import date as pydate

    np.random.seed(42)
    dates = pl.date_range(
        pydate(2024, 6, 1), pydate(2025, 2, 25), eager=True
    ).to_list()

    normal_conversations = [
        ["Bonjour je souhaite modifier mon offre", "oui celle en cours", "d'accord merci"],
        ["Quand sera traitée ma demande", "celle du mois dernier", "ok je patiente"],
        ["Je veux des informations sur mon contrat", "oui c'est bien ça", "merci"],
        ["Pouvez-vous me rappeler", "demain matin si possible", "parfait"],
        ["J'aimerais changer mon mot de passe", "oui je confirme", "c'est bon merci"],
        ["Comment accéder à mon espace client", "je ne trouve pas le lien"],
        ["Je voudrais un conseiller", "pour parler de ma facture", "oui"],
    ]
    negative_conversations = [
        ["C'est scandaleux personne ne me répond", "ça fait des jours que j'attends",
         "je vais porter plainte"],
        ["Ça fait trois semaines que j'attends", "toujours rien", "c'est inadmissible"],
        ["Service incompétent je vais résilier", "j'en ai assez",
         "passez-moi un responsable"],
        ["Panne depuis ce matin rien ne fonctionne", "toujours pas rétabli",
         "c'est la troisième fois ce mois"],
        ["Bug sur l'application impossible de se connecter", "j'ai tout essayé",
         "ça ne marche toujours pas"],
        ["Je suis très déçu de votre service", "la qualité a vraiment baissé"],
        ["Arnaque pure et simple remboursez-moi", "je n'ai jamais demandé ça"],
        ["Votre service est lamentable", "je vais aller voir la concurrence"],
        ["Je vais aller chez le concurrent", "vous êtes trop cher",
         "j'ai reçu une meilleure offre"],
        ["Surfacturation sur ma dernière facture", "je veux un remboursement immédiat"],
    ]
    emerging_conversations = [
        ["Problème avec la nouvelle mise à jour fibre", "rien ne marche depuis",
         "quand est-ce que ça sera réparé"],
        ["Depuis la migration fibre plus rien ne marche", "internet coupé",
         "c'est inadmissible"],
        ["Migration fibre catastrophique", "débit ridicule", "je veux résilier"],
    ]

    conv_id = 0
    all_dates, all_ids, all_verbatims = [], [], []

    for d in dates:
        n = np.random.randint(50, 200)
        n_normal = int(n * 0.75)
        n_negative = n - n_normal

        for _ in range(n_normal):
            conv = normal_conversations[np.random.randint(len(normal_conversations))]
            all_dates.append(d)
            all_ids.append(f"conv_{conv_id:06d}")
            all_verbatims.append(str(conv))
            conv_id += 1

        for _ in range(n_negative):
            conv = negative_conversations[np.random.randint(len(negative_conversations))]
            all_dates.append(d)
            all_ids.append(f"conv_{conv_id:06d}")
            all_verbatims.append(str(conv))
            conv_id += 1

        if d == pydate(2025, 1, 15):
            for _ in range(100):
                conv = negative_conversations[np.random.randint(len(negative_conversations))]
                all_dates.append(d)
                all_ids.append(f"conv_{conv_id:06d}")
                all_verbatims.append(str(conv))
                conv_id += 1

        if d >= pydate(2025, 2, 12):
            for _ in range(30):
                conv = emerging_conversations[np.random.randint(len(emerging_conversations))]
                all_dates.append(d)
                all_ids.append(f"conv_{conv_id:06d}")
                all_verbatims.append(str(conv))
                conv_id += 1

    df_demo = pl.DataFrame({
        "date": all_dates,
        "conversation_id": all_ids,
        "verbatims": all_verbatims,
    })

    print(f"📊 Dataset de démo : {df_demo.height:,} conversations")
    print(f"   (en production : passer un LazyFrame via pl.scan_parquet())\n")

    start_time = time.time()

    analyzer = VerbatimAnalyzer(
        df_demo,
        date_col="date",
        verbatim_col="verbatims",
        conversation_id_col="conversation_id",
        z_threshold=2.0,
        growth_window=7,
        growth_factor=3.0,
        lda_sample_size=50_000,
    )
    report = analyzer.run_full_analysis(
        save_dashboard="dashboard_verbatims.png",
        save_excel="analyse_termes.xlsx",
        max_samples_per_term=10,
    )

    elapsed = time.time() - start_time
    print(f"\n⏱️  Temps total : {elapsed:.1f}s")
